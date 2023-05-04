import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
lt=[]
#lt=np.array
s = 'ExpNumpy.xlsx'
#wb=workbook ()
wb2=load_workbook(s)
#print(wb2.sheetnames)
#ws=wb2.active
ws=wb2.get_sheet_by_name('Sheet1')
print(ws)
for columns in ws.values:
    lt+=[columns]
#print(lt)

#for columns in ws.values:
#    if lt=='':
#        lt=np.array([columns])
    #else:
#        lt+=np +([columns])
        
print(lt)
print('____________________________________________________________')
print(lt[1][1])


